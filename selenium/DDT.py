'''
Data Driven Testing is a testing methodlogy where test data is seperated from the test logic and stored externally (e
e ,g in excel , csv , json, databases) the same test script runs multiple times with different sets of data to validate scenariros.


why use data driven Testing?

1, Reduce code duplication
2, Make it easy to manage and update  test data 
3, Ideal for regrestion testing and boundary testing 


'''


from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

#load the exel file

workbook =  load_workbook("testdata.xlsx")
sheet =  workbook.active

#launch chrome browser
driver  =  webdriver.Chrome()

#loop through rows(skipping header)

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username =  row[0]
    password = row[1]


    #open the website

    driver.get("https://www.saucedemo.com/")

    #perform logic
# Perform login
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()

    # Open menu and logout
    driver.find_element(By.ID, "react-burger-menu-btn").click()
    time.sleep(5)
    driver.find_element(By.ID, "logout_sidebar_link").click()

# Close browser
driver.quit()
    